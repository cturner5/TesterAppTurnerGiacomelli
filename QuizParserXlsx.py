import openpyxl
import pprint
import re

class Quiz:
    def __init__(self):
        self.fileName = ""  # name of file to be used
        self.questions = []  # list of questions
        self.answers = []  # list of answers, answers list corresponds with the questions in the questions list
        self.numbers = []  # list of question numberts
        self.numbers_answers = [] # list of dictionaries, each element's key is a number and its value is the answer to its corresponding question
        self.quiz = {}  # dictionary made by combining questions and answers

    def parseQuestionsXlsx(self, fileName): #gets questions from the quiz xlsx document
        location = "quizFiles\\" + fileName #gets file out of quizFiles directory
        quizFile = openpyxl.load_workbook(location) #opens quiz file
        sheet_names = quizFile.sheetnames #gets sheets
        sheet = quizFile[sheet_names[0]]    #gets first sheet

        tempString = ""        #initialized temporary string for storing each question
        for cellObj in sheet['B']: #iterates through the quiz file
            if cellObj.value != None and sheet['C'][cellObj.row - 1].value == None: #gets the question
                tempString = tempString + str(cellObj.value) + "\n"
            elif cellObj.value != None and sheet['C'][cellObj.row - 1].value != None: #gets the answers
                tempString = tempString + str(cellObj.value) + " " + str(sheet['C'][cellObj.row - 1].value) + "\n"
            if cellObj.value == None:   #end of question and answers
                self.questions.append(tempString) #appends string to list of questions
                tempString = "" #resets temp string

        self.questions.append(tempString) #append last question to question list

        return self.questions   #return question list

    def parseAnswersXlsx(self, fileName): #gets answers that correspond to selected questions

        for i in range(5):  #removes .xlsx from filename
            fileName = fileName[:-1]

        location = "quizFiles\\quizAnswers\\" + fileName + "Answers.xlsx"   #gets answers file based on filename

        answerFile = openpyxl.load_workbook(location)   #open answers file
        sheet_names = answerFile.sheetnames #get sheets
        sheet = answerFile[sheet_names[0]]  #get first sheet

        for cellObj in sheet['A']:      #iterates through the file
            self.answers.append(str(cellObj.value)) #appends answers to the answers list
        #print(len(self.answers)) for testing

        samp = ""
        for i in self.answers:
            samp = samp + i

        numbersRegex = re.compile(r'\d+')

        self.numbers = [x for x in numbersRegex.findall(samp)]

        for i in range(len(self.numbers)):
            self.numbers_answers.append({self.numbers[i]: self.answers[i]})

        pprint.pprint(self.numbers_answers)

        return self.numbers_answers
        # questions_or_questionRange

    def questions_or_questionRange(self):
        # in order to ask for number/range of questions, it must first know what questions are in the quiz
        # to do this, it must find the number of the first question
        counter = 1
        numberList = []
        pprint.pprint(self.questions)
        for i in self.questions:
            numberList.append(counter)
            counter += 1
        print(numberList)
        # ^^this is the only part that is suspect, it would be BEST if it directly scans self.questions and creates the
        # numberList from the actual questions in the quiz rather than basing it on how many questions
        # are in self.questions

        # this function returns a tuple. if the user selects a number, if will output a tuple where the first element
        # is the number of questions to be randomly chosen, and the second is 0
        # if the user selects a range of questions, the first element of the tuple will be the lower bound and the
        # second element will be the upper bound

        print("Press '1' to enter a number of questions:")
        print("\t\t\t-or-")
        print("Press '2' to enter a range of questions:")
        while True:
            try:
                option = int(input())
                if option != 1 and option != 2:
                    raise ValueError
                break
            except ValueError:
                print("Invalid entry, reenter:")

        if option == 1:
            print("Enter how many questions you want on your quiz:")
            while True:
                try:
                    numbers = int(input())
                    if isinstance(numbers, int) != True or numbers > len(numberList) or numbers < 1:
                        raise ValueError
                    break
                except ValueError:
                    print("Invalid entry, reenter:")

            return (numbers, 0)

        elif option == 2:
            print("Enter the lower bound:")
            while True:
                try:
                    lower = int(input())
                    if isinstance(lower, int) != True or lower not in numberList:
                        raise ValueError
                    break
                except ValueError:
                    print("Invalid entry, reenter:")
            print("Enter the upper bound:")
            while True:
                try:
                    upper = int(input())
                    if isinstance(upper, int) != True or upper not in numberList or upper < lower:
                        raise ValueError
                    break
                except ValueError:
                    print("Invalid entry, reenter:")
            return (lower, upper)
    # questions_or_questionRange

    def parseQuizXlsx(self):
        # creates final data structure of form {question:{number:answer}}
        # needs only self.numbers_answers and self.questions

        for i in range(len(self.questions)):
            self.quiz[self.questions[i]] = self.numbers_answers[i]

        return self.quiz

# Quiz.parseQuestionsXlsx("",'textxlsx.xlsx') #for testing